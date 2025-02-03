#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pm_matrix_builder.py

Project Management Matrix Builder - A module for creating
Work Breakdown Structure (WBS) visualizations using a YAML structure.

Author: Brighton Sikarskie
Date: 2025-01-29
"""

import logging
import re
from graphviz import Digraph
import yaml
from typing import Dict, List, Union, Tuple, Any
import pandas as pd
import openpyxl

# Configure module-level logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter(
    fmt="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
if not logger.handlers:
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

def load_wbs_from_yaml(yaml_file: str) -> Dict:
    """
    Load WBS structure from a YAML file.
    
    Args:
        yaml_file: Path to the YAML file containing WBS structure
        
    Returns:
        Dictionary containing the parsed WBS structure
    """
    logger.info(f"Loading WBS structure from {yaml_file}")
    try:
        with open(yaml_file, 'r') as f:
            wbs_data = yaml.safe_load(f)
        return wbs_data
    except Exception as e:
        logger.error(f"Error loading YAML file: {e}", exc_info=True)
        raise

def is_leaf_node(item: Dict) -> bool:
    """Check if an item is a leaf node with RAM attributes."""
    return isinstance(item, dict) and 'responsibilities' in item and 'duration' in item and 'labor' in item

def convert_yaml_to_wbs_format(yaml_data: Dict) -> Dict[str, List[Tuple[str, List]]]:
    """
    Convert YAML data structure to WBS format, skipping nodes with only one child.
    
    Args:
        yaml_data: Dictionary containing the YAML data structure
        
    Returns:
        Dictionary in WBS format with (name, children) tuples
    """
    def process_item(name: str, items: Any) -> List[Tuple[str, List]]:
        if not items:
            return []
        
        # If this is a leaf node with RAM attributes, return empty list
        if is_leaf_node(items):
            return []

        result = []
        if isinstance(items, list):
            # Handle list of items
            for item in items:
                if isinstance(item, dict):
                    for child_name, child_items in item.items():
                        child_result = process_item(child_name, child_items)
                        if child_result or is_leaf_node(child_items):
                            result.append((child_name, child_result))
        elif isinstance(items, dict):
            # Handle dictionary items, excluding metadata fields
            metadata_fields = {'type', 'responsibilities', 'duration', 'labor', 'wbs_number', 'name'}
            valid_children = []
            
            # First, collect all valid children
            for child_name, child_items in items.items():
                if child_name not in metadata_fields and isinstance(child_items, dict):
                    child_result = process_item(child_name, child_items)
                    if child_result or is_leaf_node(child_items):
                        valid_children.append((child_name, child_items, child_result))
            
            # Process children based on count
            for child_name, child_items, child_result in valid_children:
                if len(valid_children) >= 2 or is_leaf_node(child_items):
                    # Keep this node if parent has multiple children or if it's a leaf
                    result.append((child_name, child_result))
                else:
                    # Skip single-child nodes by adding their children directly
                    result.extend(child_result)

        return result

    # Process the root level
    result = {}
    for root_name, root_items in yaml_data.items():
        if isinstance(root_items, dict):
            # Get project name from the name tag if it exists
            project_name = root_items.get('name', root_name)
            # Exclude metadata fields at root level
            metadata_fields = {'type', 'responsibilities', 'duration', 'labor', 'wbs_number', 'name'}
            children = []
            valid_children = []
            
            # First, collect all valid children
            for child_name, child_items in root_items.items():
                if child_name not in metadata_fields and isinstance(child_items, dict):
                    child_result = process_item(child_name, child_items)
                    if child_result or is_leaf_node(child_items):
                        valid_children.append((child_name, child_items, child_result))
            
            # Process children based on count
            for child_name, child_items, child_result in valid_children:
                if len(valid_children) >= 2 or is_leaf_node(child_items):
                    # Keep this node if parent has multiple children or if it's a leaf
                    children.append((child_name, child_result))
                else:
                    # Skip single-child nodes by adding their children directly
                    children.extend(child_result)
            
            result[project_name] = children
    
    return result

def create_wbs_from_yaml(
    yaml_file: str,
    output_filename: str = "wbs_output",
    file_format: str = "pdf",
    graph_direction: str = "TB",
    colors: dict = None
) -> None:
    """
    Create a WBS diagram from a YAML file.
    
    Args:
        yaml_file: Path to the YAML file containing WBS structure
        output_filename: Name of the output file (without extension)
        file_format: Output format (pdf, png, etc.)
        graph_direction: Direction of the graph ("TB" for top-bottom, "LR" for left-right)
        colors: Optional color scheme for different levels
    """
    yaml_data = load_wbs_from_yaml(yaml_file)
    wbs_structure = convert_yaml_to_wbs_format(yaml_data)
    create_wbs_diagram(wbs_structure, output_filename, file_format, graph_direction, colors)

class WBSColors:
    """Color schemes for WBS diagrams."""
    DEFAULT = {
        0: "#fbf1c7",  # Light0
        1: "#d3869b",  # Purple
        2: "#8ec07c",  # Green
        3: "#fe8019",  # Orange
        4: "#83a598",  # Blue
        5: "#b8bb26",  # Yellow
        6: "#fb4934",  # Red
        7: "#fabd2f"   # Bright Yellow
    }

def _safe_node_id(label: str) -> str:
    """Convert a label into a safe node ID (no spaces, punctuation)."""
    return re.sub(r"[^a-zA-Z0-9_]", "_", label)

def _node_style_for(label: str, level: int = 0, colors: dict = None) -> dict:
    """
    Returns style dict for a node.
    level: 0 for root, 1 for first level, 2 for second level, 3+ for deeper levels
    colors: Optional dictionary mapping levels to colors
    """
    if colors is None:
        colors = WBSColors.DEFAULT

    # Adjust size based on level
    sizes = {
        0: ("0.5", "0.3"),  # (width, height) for root
        1: ("0.45", "0.3"),
        2: ("0.4", "0.25"),
        3: ("0.35", "0.25"),
        4: ("0.3", "0.2"),
        5: ("0.25", "0.2"),
        6: ("0.2", "0.15"),
        7: ("0.15", "0.15")
    }
    width, height = sizes.get(level, ("0.15", "0.15"))
    
    # Adjust font size based on level
    font_sizes = {
        0: "12",
        1: "11",
        2: "10",
        3: "9",
        4: "8",
        5: "8",
        6: "7",
        7: "7"
    }

    return {
        "shape": "box",
        "style": "filled",
        "color": "#3c3836",  # Gruvbox dark0
        "fillcolor": colors.get(level % len(colors), "#fbf1c7"),
        "fontcolor": "#282828",  # Gruvbox dark
        "fontname": "Arial",
        "fontsize": font_sizes.get(level, "7"),
        "height": height,
        "width": width,
        "margin": "0.05",
        "penwidth": "1.0"
    }

def create_wbs_diagram(
    wbs_structure: dict,
    output_filename: str = "wbs_output",
    file_format: str = "pdf",
    graph_direction: str = "TB",
    colors: dict = None
) -> None:
    """
    Create and save a WBS diagram with proper styling and layout.
    
    Args:
        wbs_structure: Dictionary containing the WBS structure
        output_filename: Name of the output file (without extension)
        file_format: Output format (pdf, png, etc.)
        graph_direction: Direction of the graph ("TB" for top-bottom, "LR" for left-right)
        colors: Optional color scheme for different levels
    """
    logger.info("Creating WBS diagram...")
    try:
        dot = Digraph(comment="Work Breakdown Structure", engine='dot', format=file_format)
        dot.attr(
            rankdir=graph_direction,
            splines="line",
            nodesep="0.2",
            ranksep="0.25",
            pad="0.2",
            concentrate="true",
            compound="true",
            newrank="true",
            bgcolor="#f9f5d7",   # Gruvbox light background
            dpi="600",           # High resolution but not excessive
            fontname="Arial"     # Consistent font
        )

        def add_nodes_and_edges(items, parent_node=None, level=0, parent_number=""):
            """
            Recursively add nodes and edges to the graph.
            """
            nodes = []
            
            if isinstance(items, dict):
                # Handle the root level
                for label, children in items.items():
                    node_id = _safe_node_id(label)
                    node_style = _node_style_for(label, level, colors)
                    dot.node(node_id, label, **node_style)
                    add_nodes_and_edges(children, node_id, level + 1)
            else:
                # Handle list of tuples for other levels
                for idx, (label, children) in enumerate(items, 1):
                    # Generate node number based on level and position
                    if level == 1:
                        node_number = f"{idx}.0"
                    elif level > 1:
                        if parent_number:
                            if parent_number.endswith('.0'):
                                # Remove the .0 for cleaner numbering
                                parent_base = parent_number[:-2]
                                node_number = f"{parent_base}.{idx}"
                            else:
                                node_number = f"{parent_number}.{idx}"
                        else:
                            node_number = str(idx)
                    else:
                        node_number = ""

                    # Format the display label with the WBS number
                    if level > 0:  # Don't add number to root level
                        display_label = f"{label} ({node_number})"
                    else:
                        display_label = label
                    
                    # Create unique node ID
                    node_id = _safe_node_id(f"{label}_{node_number}")
                    
                    # Add the node
                    node_style = _node_style_for(display_label, level, colors)
                    dot.node(node_id, display_label, **node_style)
                    nodes.append(node_id)
                    
                    if parent_node:
                        dot.edge(parent_node, node_id,
                               color="#3c3836",  # Gruvbox dark0
                               arrowsize="0.6",
                               penwidth="1.0")
                    
                    if children:
                        add_nodes_and_edges(children, node_id, level + 1, node_number)
            
            return nodes

        # Start with top-level items
        for top_label, top_structure in wbs_structure.items():
            top_node_id = _safe_node_id(top_label)
            dot.node(top_node_id, top_label, **_node_style_for(top_label, 0, colors))
            add_nodes_and_edges(top_structure, top_node_id, 1)

        # Render the final diagram
        output_path = dot.render(filename=output_filename, cleanup=True)
        logger.info(f"WBS diagram saved to: {output_path}")

    except Exception as e:
        logger.error("Error while creating WBS diagram:", exc_info=True)
        raise 

class RAMColors:
    """Color schemes for RAM diagrams."""
    RESPONSIBILITY = {
        'L': "#af3a03",  # Lead - Darker orange
        'P': "#427b58",  # Participant - Even darker green
        'R': "#8f3f71",  # Reviewer - Even darker purple
        'I': "#076678",  # Input - Even darker blue
        'null': "#fbf1c7" # No involvement - Light
    }
    HEADER = "#3c3836"   # Dark header
    ALTERNATE_ROW = "#f2e5bc"  # Slightly darker than background for alternating rows

def create_ram_legend(dot: Digraph) -> None:
    """Add a legend explaining responsibility types and colors."""
    with dot.subgraph(name='cluster_legend') as legend:
        legend.attr(label='Legend', style='rounded', bgcolor='white', fontsize="12")
        
        # Add legend entries
        legend_items = {
            'L: Lead': RAMColors.RESPONSIBILITY['L'],
            'P: Participant': RAMColors.RESPONSIBILITY['P'],
            'R: Reviewer': RAMColors.RESPONSIBILITY['R'],
            'I: Input': RAMColors.RESPONSIBILITY['I'],
            'No Involvement': RAMColors.RESPONSIBILITY['null']
        }
        
        for idx, (text, color) in enumerate(legend_items.items()):
            legend.node(f'legend_{idx}', text,
                      shape='box', style='filled',
                      fillcolor=color, fontcolor="#282828",
                      fontname="Arial", fontsize="12",
                      margin="0.1")
            if idx > 0:
                legend.edge(f'legend_{idx-1}', f'legend_{idx}', style='invis')

def extract_items(data, parent_name=None, level=0, parent_wbs="", path_index=None):
    """
    Extract hierarchical items from YAML data, skipping nodes with only one child.
    
    Args:
        data: Dictionary containing the YAML data
        parent_name: Name of the parent item
        level: Current hierarchy level
        parent_wbs: Parent's WBS number
        path_index: List tracking the current path indices
    
    Returns:
        List of items with their hierarchy information
    """
    items = []
    if path_index is None:
        path_index = []

    # Filter out metadata fields at the current level
    metadata_fields = {'type', 'responsibilities', 'duration', 'labor', 'name'}
    
    def is_leaf_node(node):
        """Check if a node is a leaf node (Subtask)"""
        return isinstance(node, dict) and 'responsibilities' in node

    def has_non_metadata_children(node):
        """Check if node has any non-metadata children"""
        if not isinstance(node, dict):
            return False
        return any(k not in metadata_fields and isinstance(v, dict) for k, v in node.items())

    def generate_wbs_number(path, level):
        """Generate WBS number from path and level"""
        if level == 0:  # Project level
            return ""
        elif level == 1:  # Phase level
            return f"{path[-1]}.0"
        elif level == 2:  # Activity level
            return f"{path[1]}.{path[-1]}"  # e.g., 1.1, 1.2, etc.
        else:
            # For deeper levels, remove the first number and join the rest
            return ".".join(str(x) for x in path[1:]) if len(path) > 1 else str(path[0])

    def count_valid_children(node):
        """Count the number of valid (non-metadata) children a node has"""
        if not isinstance(node, dict):
            return 0
        return sum(1 for k, v in node.items() if k not in metadata_fields and isinstance(v, dict))

    # Get all valid items (non-metadata) at this level and sort them by structure
    valid_items = []
    for k, v in data.items():
        if k not in metadata_fields and isinstance(v, dict):
            if level == 1:  # Phase level is determined by hierarchy
                # Phases come first
                valid_items.append((0, k, v))
            elif level == 2:  # Activity level is one level below Phase
                # Activities come second
                valid_items.append((1, k, v))
            elif is_leaf_node(v):
                # Leaf nodes (Subtasks) come last
                valid_items.append((2, k, v))
            else:
                # Other nodes
                valid_items.append((1, k, v))

    # Sort by the order key and remove it
    valid_items.sort(key=lambda x: x[0])
    valid_items = [(k, v) for _, k, v in valid_items]

    for i, (name, content) in enumerate(valid_items, 1):
        if isinstance(content, dict):
            # Update the path index for this item
            if parent_wbs:
                # If we have a parent WBS number, parse it and add our index
                if parent_wbs.endswith('.0'):  # Phase level
                    current_path = [int(parent_wbs[:-2])]  # Remove .0 and convert to int
                else:
                    # Split the parent WBS number and convert all parts to integers
                    current_path = [int(x) for x in parent_wbs.split('.')]
                current_path.append(i)
            else:
                current_path = path_index + [i]
            
            # Generate WBS number for display
            wbs_number = generate_wbs_number(current_path, level)
            
            # For passing to children, use the full path string
            child_wbs = ".".join(str(x) for x in current_path)
            
            # Count valid children
            child_count = count_valid_children(content)
            
            # Determine if we should skip this node
            should_skip = child_count == 1 and not is_leaf_node(content)
            
            if not should_skip:
                # Determine type based on level and content
                if level == 0:
                    current_type = "Project"
                    display_name = content.get('name', name)
                elif level == 1:  # Phase is determined by level
                    current_type = "Phase"
                    display_name = name
                elif level == 2:  # Activity is one level below Phase
                    current_type = "Activity"
                    display_name = name
                elif is_leaf_node(content):
                    current_type = "Subtask"
                    display_name = name
                else:
                    current_type = "Task"  # Everything else is a Task
                    display_name = name
                
                # Get responsibilities and other attributes
                responsibilities = content.get('responsibilities', {})
                duration = content.get('duration', '')
                labor = content.get('labor', '')
                
                # Add the current item
                items.append({
                    'name': display_name,
                    'parent': parent_name,
                    'type': current_type,
                    'wbs_number': wbs_number,
                    'responsibilities': responsibilities,
                    'duration': duration,
                    'labor': labor,
                    'level': level
                })
            
            # Process children (excluding metadata fields)
            children_data = {k: v for k, v in content.items() if k not in metadata_fields and isinstance(v, dict)}
            if children_data:
                # If we're skipping this node, use the parent's name and level
                next_parent = parent_name if should_skip else display_name
                next_level = level if should_skip else level + 1
                items.extend(extract_items(
                    children_data,
                    next_parent,
                    next_level,
                    child_wbs,
                    current_path
                ))
    
    return items

def create_ram_diagram(
    yaml_file: str,
    output_filename: str = "ram_output",
    file_format: str = "pdf"
) -> None:
    """
    Create and save a Responsibility Assignment Matrix diagram as a simple CSV-like table.
    """
    logger.info("Creating RAM diagram...")
    try:
        # Load and process YAML data
        yaml_data = load_wbs_from_yaml(yaml_file)
        
        # Get project name from YAML
        project_name = ""
        for root_name, root_items in yaml_data.items():
            if isinstance(root_items, dict):
                project_name = root_items.get('name', root_name)
                break
        
        # Create the diagram
        dot = Digraph(comment="RAM", engine='dot', format=file_format)
        dot.attr(
            rankdir="TB",
            splines="none",
            nodesep="0.2",
            ranksep="0.3",
            pad="0.2",
            bgcolor="white"
        )

        # Create the main RAM table with project name in header
        main_table_html = f'''<
        <TABLE BORDER="1" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4">
            <TR>
                <TD BGCOLOR="lightgray" COLSPAN="11"><B>{project_name} - Responsibility Assignment Matrix</B></TD>
            </TR>
            <TR>
                <TD BGCOLOR="lightgray">WBS Element/Personnel</TD>
                <TD BGCOLOR="lightgray">Work Package Type</TD>
                <TD BGCOLOR="lightgray">Work Package</TD>
                <TD BGCOLOR="lightgray">Project Manager</TD>
                <TD BGCOLOR="lightgray">Hardware</TD>
                <TD BGCOLOR="lightgray">Software</TD>
                <TD BGCOLOR="lightgray">Testing</TD>
                <TD BGCOLOR="lightgray">Sponsor</TD>
                <TD BGCOLOR="lightgray">Other</TD>
                <TD BGCOLOR="lightgray">Work Package Duration (Days)</TD>
                <TD BGCOLOR="lightgray">Work Package Labor (Person Hours)</TD>
            </TR>
        '''

        # Extract all items
        all_items = extract_items(yaml_data)

        # Track the level for WBS colors
        current_level = 0
        prev_parent = None

        # Add data rows to main table
        for item in all_items:
            # Indent name based on level and type
            indent = "&nbsp;" * (4 * item['level'])
            name_cell = f"{indent}{item['name']}"
            
            # Reset level counter when parent changes
            if item['parent'] != prev_parent:
                if item['level'] <= current_level:
                    current_level = item['level']
                prev_parent = item['parent']
            
            # Set background color based on type and level
            if item['type'] == 'Subtask':
                bg_color = WBSColors.DEFAULT.get(item['level'], WBSColors.DEFAULT[0])
            else:
                bg_color = {
                    'Project': '#f9f5d7',  # Light beige
                    'Phase': '#ebdbb2',    # Darker beige
                    'Activity': '#d5c4a1',  # Light brown
                    'Task': '#bdae93'      # Medium brown
                }.get(item['type'], 'white')
            
            row = '<TR>'
            
            # Apply background color to all cells
            row += f'<TD BGCOLOR="{bg_color}" ALIGN="left">{name_cell}</TD>'
            row += f'<TD BGCOLOR="{bg_color}" ALIGN="left">{item["type"]}</TD>'
            
            # Work package cell (WBS number)
            row += f'<TD BGCOLOR="{bg_color}" ALIGN="left">{item["wbs_number"]}</TD>'
            
            # Role cells
            roles = ['project_manager', 'hardware', 'software', 'testing', 'sponsor', 'other']
            for role in roles:
                resp = item['responsibilities'].get(role, '')
                if item['type'] == 'Subtask' and resp:
                    resp_color = {
                        'L': '#af3a03',  # Lead - Darker orange
                        'P': '#427b58',  # Participant - Even darker green
                        'R': '#8f3f71',  # Reviewer - Even darker purple
                        'I': '#076678'   # Input - Even darker blue
                    }.get(resp, bg_color)
                else:
                    resp_color = bg_color
                row += f'<TD BGCOLOR="{resp_color}" ALIGN="center">{resp}</TD>'
            
            # Duration and labor cells
            row += f'<TD BGCOLOR="{bg_color}" ALIGN="right">{item["duration"]}</TD>'
            row += f'<TD BGCOLOR="{bg_color}" ALIGN="right">{item["labor"]}</TD>'
            row += '</TR>'
            
            main_table_html += row

        # Close the main table
        main_table_html += '</TABLE>>'
        
        # Create the main RAM table node
        dot.node('ram_table', main_table_html, shape='none')

        # Calculate totals
        total_items = len(all_items)
        total_phases = sum(1 for item in all_items if item['type'] == 'Phase')
        total_activities = sum(1 for item in all_items if item['type'] == 'Activity')
        total_tasks = sum(1 for item in all_items if item['type'] == 'Task')
        total_subtasks = sum(1 for item in all_items if item['type'] == 'Subtask')
        total_duration = sum(int(item['duration']) for item in all_items if item['duration'])
        total_labor = sum(int(item['labor']) for item in all_items if item['labor'])

        # Create a table to hold both totals and legend side by side
        bottom_table_html = '''<
        <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="20" CELLPADDING="4">
        <TR>
        <TD>
            <TABLE BORDER="1" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4">
                <TR>
                    <TD BGCOLOR="lightgray" COLSPAN="2" ALIGN="center"><B>Project Totals</B></TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Items:</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_items) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Phases:</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_phases) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Activities:</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_activities) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Tasks:</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_tasks) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Work Packages:</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_subtasks) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Duration (Days):</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_duration) + '''</TD>
                </TR>
                <TR>
                    <TD BGCOLOR="lightgray" ALIGN="right">Total Labor (Hours):</TD>
                    <TD BGCOLOR="white" ALIGN="left">&nbsp;''' + str(total_labor) + '''</TD>
                </TR>
            </TABLE>
        </TD>
        <TD>
            <TABLE BORDER="0" CELLBORDER="0" CELLSPACING="0" CELLPADDING="0">
            <TR><TD HEIGHT="40"></TD></TR>
            <TR><TD>
                <TABLE BORDER="1" CELLBORDER="1" CELLSPACING="0" CELLPADDING="4">
                    <TR>
                        <TD BGCOLOR="lightgray" COLSPAN="9" ALIGN="center"><B>Key</B></TD>
                    </TR>
                    <TR>
                        <TD BGCOLOR="#af3a03" ALIGN="left">L = Lead</TD>
                        <TD BGCOLOR="#427b58" ALIGN="left">P = Participant</TD>
                        <TD BGCOLOR="#8f3f71" ALIGN="left">R = Reviewer</TD>
                        <TD BGCOLOR="#076678" ALIGN="left">I = Input</TD>
                        <TD BGCOLOR="#f9f5d7" ALIGN="left">Project</TD>
                        <TD BGCOLOR="#ebdbb2" ALIGN="left">Phase</TD>
                        <TD BGCOLOR="#d5c4a1" ALIGN="left">Activity</TD>
                        <TD BGCOLOR="#bdae93" ALIGN="left">Task</TD>
                        <TD BGCOLOR="white" ALIGN="left">Subtask*</TD>
                    </TR>
                    <TR>
                        <TD COLSPAN="9" ALIGN="left">* Subtask rows are colored to match their corresponding level in the WBS diagram</TD>
                    </TR>
                </TABLE>
            </TD></TR>
            </TABLE>
        </TD>
        </TR>
        </TABLE>
        >'''
        dot.node('bottom_section', bottom_table_html, shape='none')

        # Create invisible edge to control layout
        dot.edge('ram_table', 'bottom_section', style='invis')

        # Render the diagram
        output_path = dot.render(filename=output_filename, cleanup=True)
        logger.info(f"RAM diagram saved to: {output_path}")

    except Exception as e:
        logger.error("Error while creating RAM diagram:", exc_info=True)
        raise

def export_to_excel(all_items: List[Dict], output_filename: str = "project_data") -> None:
    """
    Export WBS and RAM data to Excel files.
    
    Args:
        all_items: List of dictionaries containing item data
        output_filename: Base name for the output files (without extension)
    """
    logger.info("Exporting data to Excel...")
    
    # Prepare WBS data
    wbs_data = []
    for item in all_items:
        wbs_data.append({
            'WBS Number': item['wbs_number'],
            'Level': item['level'],
            'Type': item['type'],
            'Name': item['name'],
            'Parent': item['parent']
        })
    
    # Create WBS DataFrame and export
    wbs_df = pd.DataFrame(wbs_data)
    wbs_excel_file = f"{output_filename}_wbs.xlsx"
    wbs_df.to_excel(wbs_excel_file, index=False, sheet_name='WBS')
    logger.info(f"WBS data exported to: {wbs_excel_file}")
    
    # Prepare RAM data with full hierarchy
    ram_data = []
    for item in all_items:
        # Create indentation based on level
        indent = "    " * item['level']
        name = f"{indent}{item['name']}"
        
        # Prepare row data
        row_data = {
            'WBS Element/Personnel': name,
            'Work Package Type': item['type'],
            'Work Package': item['wbs_number'] if item['type'] == 'Subtask' else '',
            'Project Manager': item['responsibilities'].get('project_manager', ''),
            'Hardware': item['responsibilities'].get('hardware', ''),
            'Software': item['responsibilities'].get('software', ''),
            'Testing': item['responsibilities'].get('testing', ''),
            'Sponsor': item['responsibilities'].get('sponsor', ''),
            'Other': item['responsibilities'].get('other', ''),
            'Work Package Duration (Days)': item['duration'],
            'Work Package Labor (Person Hours)': item['labor']
        }
        ram_data.append(row_data)
    
    # Create RAM DataFrame
    ram_df = pd.DataFrame(ram_data)
    
    # Export RAM to Excel with formatting
    ram_excel_file = f"{output_filename}_ram.xlsx"
    with pd.ExcelWriter(ram_excel_file, engine='openpyxl') as writer:
        ram_df.to_excel(writer, index=False, sheet_name='RAM')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['RAM']
        
        # Define styles
        header_fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Style headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Style data rows
        for row_idx, item in enumerate(all_items, start=2):
            # Set row background based on type and level
            if item['type'] == 'Subtask':
                # Remove the '#' from the color code
                color = WBSColors.DEFAULT.get(item['level'], WBSColors.DEFAULT[0])[1:]
                fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            else:
                colors = {
                    'Project': 'F9F5D7',
                    'Phase': 'EBDBB2',
                    'Activity': 'D5C4A1',
                    'Task': 'BDAE93'
                }
                color = colors.get(item['type'], 'FFFFFF')
                fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Apply background color to all cells in the row
            for cell in worksheet[row_idx]:
                cell.fill = fill
            
            # Color responsibility cells for Subtasks
            if item['type'] == 'Subtask':
                resp_colors = {
                    'L': 'AF3A03',
                    'P': '427B58',
                    'R': '8F3F71',
                    'I': '076678'
                }
                
                # Columns D-I are responsibility columns (Project Manager to Other)
                for col_idx, role in enumerate(['project_manager', 'hardware', 'software', 'testing', 'sponsor', 'other'], start=4):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    resp = item['responsibilities'].get(role, '')
                    if resp in resp_colors:
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=resp_colors[resp],
                            end_color=resp_colors[resp],
                            fill_type='solid'
                        )
        
        # Add totals section
        total_rows = len(all_items) + 3  # Leave a blank row
        total_items = len(all_items)
        total_phases = sum(1 for item in all_items if item['type'] == 'Phase')
        total_activities = sum(1 for item in all_items if item['type'] == 'Activity')
        total_tasks = sum(1 for item in all_items if item['type'] == 'Task')
        total_subtasks = sum(1 for item in all_items if item['type'] == 'Subtask')
        total_duration = sum(int(item['duration']) for item in all_items if item['duration'])
        total_labor = sum(int(item['labor']) for item in all_items if item['labor'])
        
        # Write totals header
        worksheet.cell(row=total_rows, column=1, value='Project Totals').fill = header_fill
        worksheet.cell(row=total_rows, column=1).font = openpyxl.styles.Font(bold=True)
        worksheet.merge_cells(start_row=total_rows, start_column=1, end_row=total_rows, end_column=2)
        
        # Write totals
        totals_data = [
            ('Total Items:', total_items),
            ('Total Phases:', total_phases),
            ('Total Activities:', total_activities),
            ('Total Tasks:', total_tasks),
            ('Total Work Packages:', total_subtasks),
            ('Total Duration (Days):', total_duration),
            ('Total Labor (Hours):', total_labor)
        ]
        
        for idx, (label, value) in enumerate(totals_data):
            row = total_rows + idx + 1
            worksheet.cell(row=row, column=1, value=label).fill = header_fill
            worksheet.cell(row=row, column=2, value=value)
        
        # Add legend
        legend_row = total_rows
        legend_col = 4  # Start the legend at column D
        
        # Legend header
        worksheet.cell(row=legend_row, column=legend_col, value='Key').fill = header_fill
        worksheet.cell(row=legend_row, column=legend_col).font = openpyxl.styles.Font(bold=True)
        worksheet.merge_cells(start_row=legend_row, start_column=legend_col, end_row=legend_row, end_column=legend_col + 7)
        
        # Legend entries
        legend_entries = [
            ('L = Lead', 'AF3A03'),
            ('P = Participant', '427B58'),
            ('R = Reviewer', '8F3F71'),
            ('I = Input', '076678'),
            ('Project', 'F9F5D7'),
            ('Phase', 'EBDBB2'),
            ('Activity', 'D5C4A1'),
            ('Task', 'BDAE93'),
            ('Subtask*', 'FFFFFF')
        ]
        
        # Add legend entries in a single row
        for col_idx, (text, color) in enumerate(legend_entries):
            cell = worksheet.cell(row=legend_row + 1, column=legend_col + col_idx, value=text)
            cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Add legend note
        note = '* Subtask rows are colored to match their corresponding level in the WBS diagram'
        worksheet.cell(row=legend_row + 2, column=legend_col, value=note)
        worksheet.merge_cells(start_row=legend_row + 2, start_column=legend_col, end_row=legend_row + 2, end_column=legend_col + 7)
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    logger.info(f"RAM data exported to: {ram_excel_file}")

def create_wbs_and_ram(
    yaml_file: str,
    wbs_filename: str = "wbs_output",
    ram_filename: str = "ram_output",
    file_format: str = "pdf",
    excel_filename: str = "project_data"
) -> None:
    """
    Create both WBS and RAM diagrams from the same YAML file and export to Excel.
    
    Args:
        yaml_file: Path to the YAML file containing project structure
        wbs_filename: Name of the WBS output file (without extension)
        ram_filename: Name of the RAM output file (without extension)
        file_format: Output format for diagrams (pdf, png, etc.)
        excel_filename: Base name for Excel output files (without extension)
    """
    # Load YAML data once
    yaml_data = load_wbs_from_yaml(yaml_file)
    
    # Create WBS diagram
    create_wbs_from_yaml(yaml_file, wbs_filename, file_format)
    
    # Extract items for RAM and Excel export
    all_items = extract_items(yaml_data)
    
    # Create RAM diagram
    create_ram_diagram(yaml_file, ram_filename, file_format)
    
    # Export to Excel
    export_to_excel(all_items, excel_filename) 